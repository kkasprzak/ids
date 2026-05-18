import os
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from ids.adapters.xtb_portfolio_loader import XTBPortfolioLoader
from ids.domain.ports.portfolio import NoPortfolioAvailableError, PortfolioMalformedError
from ids.domain.timezones import WARSAW
from tests.adapters.conftest import make_xlsx

pytestmark = pytest.mark.integration


def _loader(input_dir: Path) -> XTBPortfolioLoader:
    return XTBPortfolioLoader(input_dir=input_dir, ikze_account_id="99999999")


def _export_name(as_of: str) -> str:
    return f"account_ikze_99999999_pl_xlsx_2024-12-31_{as_of}.xlsx"


def _write_export(
    input_dir: Path,
    *,
    as_of: str = "2026-05-02",
    balance: Decimal = Decimal("1"),
    equity: Decimal = Decimal("2"),
    positions: list[dict] | None = None,
) -> Path:
    path = input_dir / _export_name(as_of)
    make_xlsx(
        path,
        balance=balance,
        equity=equity,
        export_dt=datetime.fromisoformat(f"{as_of}T10:30:00"),
        positions=[] if positions is None else positions,
    )
    return path


def test_single_open_position_parsed_correctly(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    _write_export(
        input_dir,
        balance=Decimal("10000.00"),
        equity=Decimal("10200.00"),
        positions=[
            {
                "Position": 101,
                "Symbol": "VWCE.DE",
                "Open price": Decimal("99.99"),
                "Market price": Decimal("101.01"),
                "Purchase value": Decimal("999.90"),
                "Gross P/L": Decimal("10.20"),
                "SL": 0,
            }
        ],
    )

    position = _loader(input_dir).load_latest().positions[0]

    assert position.symbol == "VWCE.DE"
    assert position.open_price == Decimal("99.99")
    assert position.market_price == Decimal("101.01")
    assert position.gross_pl_pln == Decimal("10.2")
    assert position.sl is None


def test_multiple_positions_parsed_order_preserved(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    _write_export(
        input_dir,
        positions=[
            {"Position": 11, "Symbol": "AAA"},
            {"Position": 12, "Symbol": "BBB"},
        ],
    )

    snapshot = _loader(input_dir).load_latest()

    assert tuple(position.id for position in snapshot.positions) == (11, 12)
    assert tuple(position.symbol for position in snapshot.positions) == ("AAA", "BBB")


def test_empty_open_positions_returns_empty_tuple(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    _write_export(input_dir)

    snapshot = _loader(input_dir).load_latest()
    assert snapshot.positions == ()


def test_sl_set_in_xlsx_maps_to_decimal(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    _write_export(input_dir, positions=[{"SL": Decimal("88.50")}])

    snapshot = _loader(input_dir).load_latest()
    assert snapshot.positions[0].sl == Decimal("88.5")


def test_fractional_volume_preserved_as_decimal(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    _write_export(input_dir, positions=[{"Volume": Decimal("5.9113")}])

    snapshot = _loader(input_dir).load_latest()
    assert snapshot.positions[0].volume == Decimal("5.9113")


def test_naive_open_time_localized_to_warsaw(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    naive_dt = datetime(2026, 5, 2, 9, 15)
    _write_export(input_dir, positions=[{"Open time": naive_dt}])

    snapshot = _loader(input_dir).load_latest()
    assert snapshot.positions[0].open_time.tzinfo == WARSAW


def test_latest_export_selection_by_as_of_date(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    _write_export(input_dir, as_of="2026-05-02", positions=[{"Symbol": "OLDER"}])
    _write_export(input_dir, as_of="2026-05-09", positions=[{"Symbol": "NEWER"}])

    snapshot = _loader(input_dir).load_latest()
    assert snapshot.as_of_date == date(2026, 5, 9)
    assert snapshot.positions[0].symbol == "NEWER"


def test_non_ikze_file_silently_ignored(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    _write_export(input_dir, positions=[{"Symbol": "IKZE"}])
    make_xlsx(
        input_dir / "account_ikze_88888888_pl_xlsx_2024-12-31_2026-05-03.xlsx",
        balance=Decimal("1"),
        equity=Decimal("2"),
        export_dt=datetime(2026, 5, 3, 10, 30),
        positions=[{"Symbol": "OTHER"}],
    )

    snapshot = _loader(input_dir).load_latest()
    assert snapshot.positions[0].symbol == "IKZE"


def test_no_matching_files_raises_with_filename_list(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    make_xlsx(
        input_dir / "account_ikze_88888888_pl_xlsx_2024-12-31_2026-05-03.xlsx",
        balance=Decimal("1"),
        equity=Decimal("2"),
        export_dt=datetime(2026, 5, 3, 10, 30),
        positions=[],
    )

    with pytest.raises(NoPortfolioAvailableError, match="no eligible fallback XLSX files"):
        _loader(input_dir).load_latest()


def test_missing_directory_raises_with_actionable_hint(tmp_path: Path) -> None:
    input_dir = tmp_path / "missing"
    with pytest.raises(NoPortfolioAvailableError, match="mkdir -p"):
        _loader(input_dir).load_latest()


def test_missing_open_position_sheet_raises_malformed(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    path = input_dir / _export_name("2026-05-02")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    assert workbook.active is not None
    workbook.active.title = "CLOSED POSITION HISTORY"
    workbook.create_sheet("PENDING ORDERS HISTORY")
    workbook.save(path)

    with pytest.raises(PortfolioMalformedError, match="OPEN POSITION"):
        _loader(input_dir).load_latest()


def test_missing_required_columns_raises_malformed(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    path = input_dir / _export_name("2026-05-02")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "OPEN POSITION 02052026"
    sheet.cell(row=3, column=2, value=datetime(2026, 5, 2, 10, 30))
    sheet.cell(row=4, column=4, value="Balance")
    sheet.cell(row=4, column=7, value="Equity")
    sheet.cell(row=5, column=4, value=1000.0)
    sheet.cell(row=5, column=7, value=1100.0)
    sheet.cell(row=7, column=1, value="Position")
    sheet.cell(row=7, column=2, value="Symbol")
    workbook.save(path)

    with pytest.raises(PortfolioMalformedError, match="labels"):
        _loader(input_dir).load_latest()


def test_load_from_path_bypass_loads_specific_file(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    first = _write_export(input_dir, as_of="2026-05-02", positions=[{"Symbol": "FIRST"}])
    _write_export(input_dir, as_of="2026-05-09", positions=[{"Symbol": "SECOND"}])

    snapshot = _loader(input_dir).load_from_path(first)
    assert snapshot.as_of_date == date(2026, 5, 2)
    assert snapshot.positions[0].symbol == "FIRST"


def test_load_from_path_non_matching_filename_uses_export_datetime(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    path = input_dir / "custom.xlsx"
    make_xlsx(
        path,
        balance=Decimal("1"),
        equity=Decimal("2"),
        export_dt=datetime(2026, 5, 2, 10, 30),
        positions=[],
    )

    snapshot = _loader(input_dir).load_from_path(path)
    assert snapshot.as_of_date == date(2026, 5, 2)


def test_load_from_path_rejects_file_from_other_ikze_account(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    path = input_dir / "account_ikze_88888888_pl_xlsx_2024-12-31_2026-05-02.xlsx"
    make_xlsx(
        path,
        balance=Decimal("1"),
        equity=Decimal("2"),
        export_dt=datetime(2026, 5, 2, 10, 30),
        positions=[],
    )

    with pytest.raises(PortfolioMalformedError, match="clearly belongs to IKZE account"):
        _loader(input_dir).load_from_path(path)


def test_load_latest_falls_back_to_mtime_and_workbook_export_datetime(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    first = input_dir / "renamed_a.xlsx"
    second = input_dir / "renamed_b.xlsx"
    make_xlsx(
        first,
        balance=Decimal("1"),
        equity=Decimal("2"),
        export_dt=datetime(2026, 5, 1, 10, 30),
        positions=[{"Symbol": "OLDER"}],
    )
    make_xlsx(
        second,
        balance=Decimal("1"),
        equity=Decimal("2"),
        export_dt=datetime(2026, 5, 3, 10, 30),
        positions=[{"Symbol": "NEWER"}],
    )
    first_mtime = datetime(2026, 5, 5, 10, 0).timestamp()
    second_mtime = datetime(2026, 5, 6, 10, 0).timestamp()
    os.utime(first, (first_mtime, first_mtime))
    os.utime(second, (second_mtime, second_mtime))

    snapshot = _loader(input_dir).load_latest()
    assert snapshot.positions[0].symbol == "NEWER"
    assert snapshot.as_of_date == date(2026, 5, 3)
