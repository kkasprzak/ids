from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook

_HEADER = [
    "Position",
    "Symbol",
    "Type",
    "Volume",
    "Open time",
    "Open price",
    "Market price",
    "Purchase value",
    "SL",
    "TP",
    "Margin",
    "Commission",
    "Swap",
    "Rollover",
    "Gross P/L",
    "Comment",
]

_CLOSED_HEADER = [
    "Position",
    "Symbol",
    "Type",
    "Volume",
    "Open time",
    "Open price",
    "Close time",
    "Close price",
    "Open origin",
    "Close origin",
    "Purchase value",
    "Sale value",
    "SL",
    "TP",
    "Margin",
    "Commission",
    "Swap",
    "Rollover",
    "Gross P/L",
    "Comment",
]

# Matches real XTB layout: closed-position header lives at row 11.
_CLOSED_HEADER_ROW = 11


def make_xlsx(  # noqa: PLR0913
    path: Path,
    *,
    balance: Decimal,
    equity: Decimal,
    export_dt: datetime,
    positions: list[dict],
    closed_positions: list[dict] | None = None,
) -> None:
    """Write a minimal XTB-like workbook for adapter tests."""
    workbook = Workbook()
    default = workbook.active
    assert default is not None
    workbook.remove(default)

    closed_sheet = workbook.create_sheet("CLOSED POSITION HISTORY")
    open_sheet = workbook.create_sheet(f"OPEN POSITION {export_dt.strftime('%d%m%Y')}")
    workbook.create_sheet("PENDING ORDERS HISTORY")
    workbook.create_sheet("CASH OPERATION HISTORY")

    open_sheet.cell(row=3, column=2, value=export_dt)
    open_sheet.cell(row=4, column=4, value="Balance")
    open_sheet.cell(row=4, column=7, value="Equity")
    open_sheet.cell(row=5, column=4, value=float(balance))
    open_sheet.cell(row=5, column=7, value=float(equity))

    for col, name in enumerate(_HEADER, start=1):
        open_sheet.cell(row=7, column=col, value=name)

    for idx, position in enumerate(positions, start=8):
        row = {
            "Position": position.get("Position", 1000 + idx),
            "Symbol": position.get("Symbol", "TEST.PL"),
            "Type": position.get("Type", "BUY"),
            "Volume": position.get("Volume", Decimal("1")),
            "Open time": position.get("Open time", export_dt),
            "Open price": position.get("Open price", Decimal("100.00")),
            "Market price": position.get("Market price", Decimal("100.00")),
            "Purchase value": position.get("Purchase value", Decimal("100.00")),
            "SL": position.get("SL", 0),
            "TP": position.get("TP", 0),
            "Margin": position.get("Margin", 0),
            "Commission": position.get("Commission", 0),
            "Swap": position.get("Swap", 0),
            "Rollover": position.get("Rollover", 0),
            "Gross P/L": position.get("Gross P/L", Decimal("0.00")),
            "Comment": position.get("Comment", ""),
        }
        for col, name in enumerate(_HEADER, start=1):
            value = row[name]
            if isinstance(value, Decimal):
                value = float(value)
            open_sheet.cell(row=idx, column=col, value=value)

    open_sheet.cell(row=8 + len(positions), column=1, value="Total")

    if closed_positions is not None:
        for col, name in enumerate(_CLOSED_HEADER, start=1):
            closed_sheet.cell(row=_CLOSED_HEADER_ROW, column=col, value=name)

        for idx, position in enumerate(closed_positions, start=_CLOSED_HEADER_ROW + 1):
            row_data = {
                "Position": position.get("Position", 2000 + idx),
                "Symbol": position.get("Symbol", "TEST.PL"),
                "Type": position.get("Type", "BUY"),
                "Volume": position.get("Volume", Decimal("1")),
                "Open time": position.get("Open time", export_dt),
                "Open price": position.get("Open price", Decimal("100.00")),
                "Close time": position.get("Close time", export_dt),
                "Close price": position.get("Close price", Decimal("110.00")),
                "Open origin": position.get("Open origin", ""),
                "Close origin": position.get("Close origin", ""),
                "Purchase value": position.get("Purchase value", Decimal("100.00")),
                "Sale value": position.get("Sale value", Decimal("110.00")),
                "SL": position.get("SL", 0),
                "TP": position.get("TP", 0),
                "Margin": position.get("Margin", 0),
                "Commission": position.get("Commission", 0),
                "Swap": position.get("Swap", 0),
                "Rollover": position.get("Rollover", 0),
                "Gross P/L": position.get("Gross P/L", Decimal("10.00")),
                "Comment": position.get("Comment", ""),
            }
            for col, name in enumerate(_CLOSED_HEADER, start=1):
                value = row_data[name]
                if isinstance(value, Decimal):
                    value = float(value)
                closed_sheet.cell(row=idx, column=col, value=value)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
