import os
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from ids.application.ports.portfolio import NoPortfolioAvailableError, PortfolioMalformedError
from ids.domain.timezones import WARSAW
from ids.infrastructure.adapters.xtb_portfolio_loader import XTBPortfolioLoader
from tests.infrastructure.adapters.conftest import (
    _CLOSED_HEADER,
    _CLOSED_HEADER_ROW,
    _HEADER,
    make_xlsx,
)

pytestmark = pytest.mark.integration


def _loader(input_dir: Path) -> XTBPortfolioLoader:
    return XTBPortfolioLoader(input_dir=input_dir, ikze_account_id="99999999")


def _export_name(as_of: str) -> str:
    return f"account_ikze_99999999_pl_xlsx_2024-12-31_{as_of}.xlsx"


def _write_export(  # noqa: PLR0913
    input_dir: Path,
    *,
    as_of: str = "2026-05-02",
    balance: Decimal = Decimal("1"),
    equity: Decimal = Decimal("2"),
    positions: list[dict] | None = None,
    closed_positions: list[dict] | None = None,
) -> Path:
    path = input_dir / _export_name(as_of)
    make_xlsx(
        path,
        balance=balance,
        equity=equity,
        export_dt=datetime.fromisoformat(f"{as_of}T10:30:00"),
        positions=[] if positions is None else positions,
        closed_positions=closed_positions,
    )
    return path


def test_single_open_position_parsed_correctly(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    _write_export(
        input_dir,
        balance=Decimal("10000.00"),
        equity=Decimal("10200.00"),
        positions=[
            {
                "Position": 101,
                "Symbol": "VWCE.DE",
                "Open price": Decimal("99.99"),
                "Market price": Decimal("101.01"),
                "Purchase value": Decimal("999.90"),
                "Gross P/L": Decimal("10.20"),
                "SL": 0,
            }
        ],
    )

    position = _loader(input_dir).load_latest().positions[0]

    assert position.symbol == "VWCE.DE"
    assert position.open_price == Decimal("99.99")
    assert position.market_price == Decimal("101.01")
    assert position.gross_pl_pln == Decimal("10.2")
    assert position.sl is None


def test_multiple_positions_parsed_order_preserved(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    _write_export(
        input_dir,
        positions=[
            {"Position": 11, "Symbol": "AAA"},
            {"Position": 12, "Symbol": "BBB"},
        ],
    )

    snapshot = _loader(input_dir).load_latest()

    assert tuple(position.id for position in snapshot.positions) == (11, 12)
    assert tuple(position.symbol for position in snapshot.positions) == ("AAA", "BBB")


def test_empty_open_positions_returns_empty_tuple(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    _write_export(input_dir)

    snapshot = _loader(input_dir).load_latest()
    assert snapshot.positions == ()


def test_sl_set_in_xlsx_maps_to_decimal(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    _write_export(input_dir, positions=[{"SL": Decimal("88.50")}])

    snapshot = _loader(input_dir).load_latest()
    assert snapshot.positions[0].sl == Decimal("88.5")


@pytest.mark.parametrize(
    ("sl_value", "expected"),
    [
        (0, None),
        (0.0, None),
        (Decimal("0"), None),
        ("0", None),
        ("0.0", None),
        ("", None),
        ("   ", None),
        (Decimal("101.25"), Decimal("101.25")),
        ("101.25", Decimal("101.25")),
    ],
)
def test_stop_loss_normalization_variants(
    tmp_path: Path, sl_value: int | float | Decimal | str, expected: Decimal | None
) -> None:
    input_dir = tmp_path / "inputs"
    _write_export(input_dir, positions=[{"SL": sl_value}])

    snapshot = _loader(input_dir).load_latest()
    assert snapshot.positions[0].sl == expected


def test_fractional_volume_preserved_as_decimal(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    _write_export(input_dir, positions=[{"Volume": Decimal("5.9113")}])

    snapshot = _loader(input_dir).load_latest()
    assert snapshot.positions[0].volume == Decimal("5.9113")


def test_naive_open_time_localized_to_warsaw(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    naive_dt = datetime(2026, 5, 2, 9, 15)
    _write_export(input_dir, positions=[{"Open time": naive_dt}])

    snapshot = _loader(input_dir).load_latest()
    assert snapshot.positions[0].open_time.tzinfo == WARSAW


def test_latest_export_selection_by_as_of_date(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    _write_export(input_dir, as_of="2026-05-02", positions=[{"Symbol": "OLDER"}])
    _write_export(input_dir, as_of="2026-05-09", positions=[{"Symbol": "NEWER"}])

    snapshot = _loader(input_dir).load_latest()
    assert snapshot.as_of_date == date(2026, 5, 9)
    assert snapshot.positions[0].symbol == "NEWER"


def test_non_ikze_file_silently_ignored(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    _write_export(input_dir, positions=[{"Symbol": "IKZE"}])
    make_xlsx(
        input_dir / "account_ikze_88888888_pl_xlsx_2024-12-31_2026-05-03.xlsx",
        balance=Decimal("1"),
        equity=Decimal("2"),
        export_dt=datetime(2026, 5, 3, 10, 30),
        positions=[{"Symbol": "OTHER"}],
    )

    snapshot = _loader(input_dir).load_latest()
    assert snapshot.positions[0].symbol == "IKZE"


def test_no_matching_files_raises_with_filename_list(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    make_xlsx(
        input_dir / "account_ikze_88888888_pl_xlsx_2024-12-31_2026-05-03.xlsx",
        balance=Decimal("1"),
        equity=Decimal("2"),
        export_dt=datetime(2026, 5, 3, 10, 30),
        positions=[],
    )

    with pytest.raises(NoPortfolioAvailableError, match="no eligible fallback XLSX files"):
        _loader(input_dir).load_latest()


def test_missing_directory_raises_with_actionable_hint(tmp_path: Path) -> None:
    input_dir = tmp_path / "missing"
    with pytest.raises(NoPortfolioAvailableError, match="mkdir -p"):
        _loader(input_dir).load_latest()


def test_missing_open_position_sheet_raises_malformed(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    path = input_dir / _export_name("2026-05-02")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    assert workbook.active is not None
    workbook.active.title = "CLOSED POSITION HISTORY"
    workbook.create_sheet("PENDING ORDERS HISTORY")
    workbook.save(path)

    with pytest.raises(PortfolioMalformedError, match="OPEN POSITION"):
        _loader(input_dir).load_latest()


def test_non_standard_sheet_name_loads_via_semantic_fallback(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    path = input_dir / _export_name("2026-05-02")
    make_xlsx(
        path,
        balance=Decimal("100.00"),
        equity=Decimal("200.00"),
        export_dt=datetime(2026, 5, 2, 10, 30),
        positions=[{"Symbol": "FALLBACK"}],
    )
    wb = load_workbook(path)
    for ws in wb.worksheets:
        if ws.title.startswith("OPEN POSITION"):
            ws.title = "POZYCJE OTWARTE 02052026"
    wb.save(path)

    snapshot = _loader(input_dir).load_latest()
    assert snapshot.positions[0].symbol == "FALLBACK"


def test_multiple_sheets_with_position_table_raises_ambiguity(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    path = input_dir / _export_name("2026-05-02")
    make_xlsx(
        path,
        balance=Decimal("1"),
        equity=Decimal("2"),
        export_dt=datetime(2026, 5, 2, 10, 30),
        positions=[{"Symbol": "ONE"}],
    )
    wb = load_workbook(path)
    for ws in wb.worksheets:
        if ws.title.startswith("OPEN POSITION"):
            ws.title = "TABLE A"
    dup = wb.create_sheet("TABLE B")
    for col, name in enumerate(_HEADER, start=1):
        dup.cell(row=1, column=col, value=name)
    wb.save(path)

    with pytest.raises(PortfolioMalformedError, match="Ambiguous"):
        _loader(input_dir).load_latest()


def test_missing_required_columns_raises_malformed(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    path = input_dir / _export_name("2026-05-02")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "OPEN POSITION 02052026"
    sheet.cell(row=3, column=2, value=datetime(2026, 5, 2, 10, 30))
    sheet.cell(row=4, column=4, value="Balance")
    sheet.cell(row=4, column=7, value="Equity")
    sheet.cell(row=5, column=4, value=1000.0)
    sheet.cell(row=5, column=7, value=1100.0)
    sheet.cell(row=7, column=1, value="Position")
    sheet.cell(row=7, column=2, value="Symbol")
    workbook.save(path)

    with pytest.raises(PortfolioMalformedError, match="required columns"):
        _loader(input_dir).load_latest()


def test_load_from_path_bypass_loads_specific_file(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    first = _write_export(input_dir, as_of="2026-05-02", positions=[{"Symbol": "FIRST"}])
    _write_export(input_dir, as_of="2026-05-09", positions=[{"Symbol": "SECOND"}])

    snapshot = _loader(input_dir).load_from_path(first)
    assert snapshot.as_of_date == date(2026, 5, 2)
    assert snapshot.positions[0].symbol == "FIRST"


def test_load_from_path_non_matching_filename_uses_export_datetime(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    path = input_dir / "custom.xlsx"
    make_xlsx(
        path,
        balance=Decimal("1"),
        equity=Decimal("2"),
        export_dt=datetime(2026, 5, 2, 10, 30),
        positions=[],
    )

    snapshot = _loader(input_dir).load_from_path(path)
    assert snapshot.as_of_date == date(2026, 5, 2)


def test_load_from_path_rejects_file_from_other_ikze_account(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    path = input_dir / "account_ikze_88888888_pl_xlsx_2024-12-31_2026-05-02.xlsx"
    make_xlsx(
        path,
        balance=Decimal("1"),
        equity=Decimal("2"),
        export_dt=datetime(2026, 5, 2, 10, 30),
        positions=[],
    )

    with pytest.raises(PortfolioMalformedError, match="clearly belongs to IKZE account"):
        _loader(input_dir).load_from_path(path)


def test_load_latest_falls_back_to_mtime_and_workbook_export_datetime(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    first = input_dir / "renamed_a.xlsx"
    second = input_dir / "renamed_b.xlsx"
    make_xlsx(
        first,
        balance=Decimal("1"),
        equity=Decimal("2"),
        export_dt=datetime(2026, 5, 1, 10, 30),
        positions=[{"Symbol": "OLDER"}],
    )
    make_xlsx(
        second,
        balance=Decimal("1"),
        equity=Decimal("2"),
        export_dt=datetime(2026, 5, 3, 10, 30),
        positions=[{"Symbol": "NEWER"}],
    )
    first_mtime = datetime(2026, 5, 5, 10, 0).timestamp()
    second_mtime = datetime(2026, 5, 6, 10, 0).timestamp()
    os.utime(first, (first_mtime, first_mtime))
    os.utime(second, (second_mtime, second_mtime))

    snapshot = _loader(input_dir).load_latest()
    assert snapshot.positions[0].symbol == "NEWER"
    assert snapshot.as_of_date == date(2026, 5, 3)


def _make_xlsx_with_row_offset(
    path: Path,
    *,
    export_dt_row: int,
    account_label_row: int,
    position_header_row: int,
) -> None:
    """Write a workbook where export datetime, account labels, and position header
    are placed at explicitly shifted row positions to test dynamic discovery."""
    wb = Workbook()
    default = wb.active
    assert default is not None
    wb.remove(default)
    ws = wb.create_sheet("OPEN POSITION 02052026")

    ws.cell(row=export_dt_row, column=2, value=datetime(2026, 5, 2, 10, 30))
    ws.cell(row=account_label_row, column=4, value="Balance")
    ws.cell(row=account_label_row, column=7, value="Equity")
    ws.cell(row=account_label_row + 1, column=4, value=1000.0)
    ws.cell(row=account_label_row + 1, column=7, value=2000.0)
    for col, name in enumerate(_HEADER, start=1):
        ws.cell(row=position_header_row, column=col, value=name)
    ws.cell(row=position_header_row + 1, column=1, value=101)
    ws.cell(row=position_header_row + 1, column=2, value="SHIFTED")
    ws.cell(row=position_header_row + 1, column=3, value="BUY")
    ws.cell(row=position_header_row + 1, column=4, value=1.0)
    ws.cell(row=position_header_row + 1, column=5, value=datetime(2026, 5, 1, 9, 0))
    ws.cell(row=position_header_row + 1, column=6, value=100.0)
    ws.cell(row=position_header_row + 1, column=7, value=110.0)
    ws.cell(row=position_header_row + 1, column=8, value=100.0)
    ws.cell(row=position_header_row + 1, column=9, value=0)
    ws.cell(row=position_header_row + 1, column=15, value=10.0)
    ws.cell(row=position_header_row + 2, column=1, value="Total")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def test_account_labels_beyond_original_row_10_are_found(tmp_path: Path) -> None:
    path = tmp_path / "inputs" / _export_name("2026-05-02")
    _make_xlsx_with_row_offset(path, export_dt_row=2, account_label_row=13, position_header_row=15)
    snapshot = _loader(tmp_path / "inputs").load_latest()
    assert snapshot.account.balance_pln == Decimal("1000")
    assert snapshot.account.equity_pln == Decimal("2000")


def test_position_header_beyond_original_row_20_is_found(tmp_path: Path) -> None:
    path = tmp_path / "inputs" / _export_name("2026-05-02")
    _make_xlsx_with_row_offset(path, export_dt_row=2, account_label_row=5, position_header_row=22)
    snapshot = _loader(tmp_path / "inputs").load_latest()
    assert snapshot.positions[0].symbol == "SHIFTED"


def test_export_datetime_beyond_original_row_6_is_found(tmp_path: Path) -> None:
    path = tmp_path / "inputs" / _export_name("2026-05-02")
    _make_xlsx_with_row_offset(path, export_dt_row=10, account_label_row=12, position_header_row=14)
    snapshot = _loader(tmp_path / "inputs").load_latest()
    assert snapshot.account.export_datetime.date() == date(2026, 5, 2)


def _make_xlsx_with_custom_headers(path: Path, headers: list[str]) -> None:
    """Write a workbook with the given header row instead of the standard one."""
    make_xlsx(
        path,
        balance=Decimal("1"),
        equity=Decimal("2"),
        export_dt=datetime(2026, 5, 2, 10, 30),
        positions=[{"Symbol": "TEST"}],
    )
    wb = load_workbook(path)
    ws = wb["OPEN POSITION 02052026"]
    for col, name in enumerate(headers, start=1):
        ws.cell(row=7, column=col, value=name)
    wb.save(path)


def test_column_labels_with_extra_whitespace_are_matched(tmp_path: Path) -> None:
    path = tmp_path / "inputs" / _export_name("2026-05-02")
    padded = [f"  {h}  " if h else h for h in _HEADER]
    _make_xlsx_with_custom_headers(path, padded)

    snapshot = _loader(tmp_path / "inputs").load_latest()
    assert snapshot.positions[0].symbol == "TEST"


def test_column_labels_with_different_case_are_matched(tmp_path: Path) -> None:
    path = tmp_path / "inputs" / _export_name("2026-05-02")
    uppercased = [h.upper() for h in _HEADER]
    _make_xlsx_with_custom_headers(path, uppercased)

    snapshot = _loader(tmp_path / "inputs").load_latest()
    assert snapshot.positions[0].symbol == "TEST"


def test_gross_pl_ampersand_alias_is_accepted(tmp_path: Path) -> None:
    path = tmp_path / "inputs" / _export_name("2026-05-02")
    alt_headers = [h if h != "Gross P/L" else "Gross P&L" for h in _HEADER]
    _make_xlsx_with_custom_headers(path, alt_headers)

    snapshot = _loader(tmp_path / "inputs").load_latest()
    assert snapshot.positions[0].gross_pl_pln == Decimal("0")


def test_missing_required_logical_field_names_it_in_error(tmp_path: Path) -> None:
    path = tmp_path / "inputs" / _export_name("2026-05-02")
    # Remove "Gross P/L" entirely from the header
    stripped = [h if h != "Gross P/L" else "" for h in _HEADER]
    _make_xlsx_with_custom_headers(path, stripped)

    with pytest.raises(PortfolioMalformedError, match="gross_pl"):
        _loader(tmp_path / "inputs").load_latest()


def _make_xlsx_with_custom_footer(path: Path, *, footer: str, n_positions: int = 2) -> None:
    """Write a workbook replacing the standard 'Total' footer with a custom string."""
    make_xlsx(
        path,
        balance=Decimal("1"),
        equity=Decimal("2"),
        export_dt=datetime(2026, 5, 2, 10, 30),
        positions=[{"Symbol": f"POS{i}"} for i in range(n_positions)],
    )
    wb = load_workbook(path)
    ws = wb["OPEN POSITION 02052026"]
    footer_row = 8 + n_positions
    ws.cell(row=footer_row, column=1, value=footer)
    wb.save(path)


def test_uppercase_total_footer_stops_parsing(tmp_path: Path) -> None:
    path = tmp_path / "inputs" / _export_name("2026-05-02")
    _make_xlsx_with_custom_footer(path, footer="TOTAL", n_positions=2)

    snapshot = _loader(tmp_path / "inputs").load_latest()
    assert len(snapshot.positions) == 2  # noqa: PLR2004


def test_footer_with_surrounding_spaces_stops_parsing(tmp_path: Path) -> None:
    path = tmp_path / "inputs" / _export_name("2026-05-02")
    _make_xlsx_with_custom_footer(path, footer="  Total  ", n_positions=1)

    snapshot = _loader(tmp_path / "inputs").load_latest()
    assert len(snapshot.positions) == 1


def test_malformed_position_row_with_numeric_id_raises_error(tmp_path: Path) -> None:
    path = tmp_path / "inputs" / _export_name("2026-05-02")
    make_xlsx(
        path,
        balance=Decimal("1"),
        equity=Decimal("2"),
        export_dt=datetime(2026, 5, 2, 10, 30),
        positions=[{"Symbol": "GOOD"}],
    )
    wb = load_workbook(path)
    ws = wb["OPEN POSITION 02052026"]
    # Row 9 is the second data row (after header row 7 and first position row 8).
    # Write a numeric position_id so footer detection passes, but corrupt open_price.
    for col, _name in enumerate(_HEADER, start=1):
        ws.cell(row=9, column=col, value=None)
    ws.cell(row=9, column=1, value=999)  # numeric position_id → not a footer
    ws.cell(row=9, column=2, value="BAD_SYMBOL")
    ws.cell(row=9, column=3, value="BUY")
    ws.cell(row=9, column=4, value=1.0)
    ws.cell(row=9, column=5, value=datetime(2026, 5, 1, 9, 0))
    ws.cell(row=9, column=6, value="NOT_A_NUMBER")  # open_price — will fail Decimal conversion
    ws.cell(row=9, column=7, value=100.0)
    ws.cell(row=9, column=8, value=100.0)
    ws.cell(row=9, column=9, value=0)
    ws.cell(row=9, column=15, value=0.0)
    wb.save(path)

    with pytest.raises(PortfolioMalformedError, match=r"row 9, field `open_price`"):
        _loader(tmp_path / "inputs").load_latest()


def test_invalid_position_type_reports_row_and_field(tmp_path: Path) -> None:
    path = tmp_path / "inputs" / _export_name("2026-05-02")
    make_xlsx(
        path,
        balance=Decimal("1"),
        equity=Decimal("2"),
        export_dt=datetime(2026, 5, 2, 10, 30),
        positions=[{"Type": "NOT_A_POSITION_TYPE"}],
    )

    with pytest.raises(PortfolioMalformedError, match=r"row 8, field `type`"):
        _loader(tmp_path / "inputs").load_latest()


def test_invalid_open_time_reports_row_and_field(tmp_path: Path) -> None:
    path = tmp_path / "inputs" / _export_name("2026-05-02")
    make_xlsx(
        path,
        balance=Decimal("1"),
        equity=Decimal("2"),
        export_dt=datetime(2026, 5, 2, 10, 30),
        positions=[{"Open time": "NOT_A_DATETIME"}],
    )

    with pytest.raises(PortfolioMalformedError, match=r"row 8, field `open_time`"):
        _loader(tmp_path / "inputs").load_latest()


def test_missing_numeric_cell_reports_row_and_field(tmp_path: Path) -> None:
    path = tmp_path / "inputs" / _export_name("2026-05-02")
    make_xlsx(
        path,
        balance=Decimal("1"),
        equity=Decimal("2"),
        export_dt=datetime(2026, 5, 2, 10, 30),
        positions=[{"Open price": None}],
    )

    with pytest.raises(PortfolioMalformedError, match=r"row 8, field `open_price`"):
        _loader(tmp_path / "inputs").load_latest()


# ---------------------------------------------------------------------------
# Closed positions
# ---------------------------------------------------------------------------


def test_single_closed_position_parsed_correctly(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    _write_export(
        input_dir,
        closed_positions=[
            {
                "Position": 501,
                "Symbol": "VWCE.DE",
                "Type": "BUY",
                "Volume": Decimal("2"),
                "Open time": datetime(2026, 1, 10, 9, 0),
                "Open price": Decimal("95.00"),
                "Close time": datetime(2026, 5, 1, 15, 0),
                "Close price": Decimal("102.50"),
                "Purchase value": Decimal("190.00"),
                "Gross P/L": Decimal("15.00"),
            }
        ],
    )

    snapshot = _loader(input_dir).load_latest()

    assert len(snapshot.closed_positions) == 1
    cp = snapshot.closed_positions[0]
    assert cp.id == 501  # noqa: PLR2004
    assert cp.symbol == "VWCE.DE"
    assert cp.open_price == Decimal("95.00")
    assert cp.close_price == Decimal("102.50")
    assert cp.purchase_value_pln == Decimal("190.00")
    assert cp.gross_pl_pln == Decimal("15.00")
    assert cp.open_time.tzinfo == WARSAW
    assert cp.close_time.tzinfo == WARSAW


def test_mix_open_and_closed_positions(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    _write_export(
        input_dir,
        positions=[{"Position": 11, "Symbol": "OPEN1"}],
        closed_positions=[
            {"Position": 501, "Symbol": "CLOSED1"},
            {"Position": 502, "Symbol": "CLOSED2"},
        ],
    )

    snapshot = _loader(input_dir).load_latest()

    assert len(snapshot.positions) == 1
    assert snapshot.positions[0].symbol == "OPEN1"
    assert len(snapshot.closed_positions) == 2  # noqa: PLR2004
    assert tuple(cp.symbol for cp in snapshot.closed_positions) == ("CLOSED1", "CLOSED2")


def test_empty_closed_sheet_yields_empty_tuple(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    _write_export(input_dir)

    snapshot = _loader(input_dir).load_latest()
    assert snapshot.closed_positions == ()


def test_malformed_closed_position_row_raises_error(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    path = input_dir / _export_name("2026-05-02")
    export_dt = datetime(2026, 5, 2, 10, 30)
    make_xlsx(
        path,
        balance=Decimal("1"),
        equity=Decimal("2"),
        export_dt=export_dt,
        positions=[],
        closed_positions=[{"Close price": "NOT_A_NUMBER"}],
    )

    with pytest.raises(
        PortfolioMalformedError, match=r"closed-position row 12, field `close_price`"
    ):
        _loader(input_dir).load_latest()


def _make_xlsx_with_closed_header_at_row(
    path: Path,
    *,
    closed_header_row: int,
    extra_cells_in_scan_window: bool = False,
) -> None:
    """Write a workbook where the closed-position header is at a custom row.

    If extra_cells_in_scan_window is True, writes a dummy cell at row 1 of the
    CLOSED sheet so the empty-sheet check (rows 1..15) sees content even when the
    header is beyond row 15.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    export_dt = datetime(2026, 5, 2, 10, 30)
    wb = Workbook()
    default = wb.active
    assert default is not None
    wb.remove(default)

    closed_ws = wb.create_sheet("CLOSED POSITION HISTORY")
    if extra_cells_in_scan_window:
        closed_ws.cell(row=1, column=1, value="DUMMY")
    for col, name in enumerate(_CLOSED_HEADER, start=1):
        closed_ws.cell(row=closed_header_row, column=col, value=name)
    closed_ws.cell(row=closed_header_row + 1, column=1, value=888)
    closed_ws.cell(row=closed_header_row + 1, column=2, value="OFFSET")
    closed_ws.cell(row=closed_header_row + 1, column=3, value="BUY")
    closed_ws.cell(row=closed_header_row + 1, column=4, value=1.0)
    closed_ws.cell(row=closed_header_row + 1, column=5, value=export_dt)
    closed_ws.cell(row=closed_header_row + 1, column=6, value=100.0)
    closed_ws.cell(row=closed_header_row + 1, column=7, value=export_dt)
    closed_ws.cell(row=closed_header_row + 1, column=8, value=110.0)
    closed_ws.cell(row=closed_header_row + 1, column=11, value=100.0)
    closed_ws.cell(row=closed_header_row + 1, column=19, value=10.0)

    open_ws = wb.create_sheet(f"OPEN POSITION {export_dt.strftime('%d%m%Y')}")
    open_ws.cell(row=3, column=2, value=export_dt)
    open_ws.cell(row=4, column=4, value="Balance")
    open_ws.cell(row=4, column=7, value="Equity")
    open_ws.cell(row=5, column=4, value=1000.0)
    open_ws.cell(row=5, column=7, value=2000.0)
    for col, name in enumerate(_HEADER, start=1):
        open_ws.cell(row=7, column=col, value=name)
    open_ws.cell(row=8, column=1, value="Total")
    wb.save(path)


def test_closed_header_at_row_5_found(tmp_path: Path) -> None:
    path = tmp_path / "inputs" / _export_name("2026-05-02")
    _make_xlsx_with_closed_header_at_row(path, closed_header_row=5)

    snapshot = _loader(tmp_path / "inputs").load_latest()
    assert snapshot.closed_positions[0].symbol == "OFFSET"


def test_closed_header_at_row_15_found(tmp_path: Path) -> None:
    """Header at the last row of the scan window (15) must still be discovered."""
    path = tmp_path / "inputs" / _export_name("2026-05-02")
    _make_xlsx_with_closed_header_at_row(path, closed_header_row=15)

    snapshot = _loader(tmp_path / "inputs").load_latest()
    assert snapshot.closed_positions[0].symbol == "OFFSET"


def test_closed_header_at_row_16_with_content_raises_malformed(tmp_path: Path) -> None:
    """Header beyond scan window (row 16) with cells in rows 1-15 → PortfolioMalformedError."""
    path = tmp_path / "inputs" / _export_name("2026-05-02")
    _make_xlsx_with_closed_header_at_row(
        path, closed_header_row=16, extra_cells_in_scan_window=True
    )

    with pytest.raises(PortfolioMalformedError, match="required columns"):
        _loader(tmp_path / "inputs").load_latest()


def test_truly_empty_closed_sheet_yields_empty_tuple(tmp_path: Path) -> None:
    """CLOSED POSITION HISTORY sheet exists but has no cells → returns ()."""
    input_dir = tmp_path / "inputs"
    path = input_dir / _export_name("2026-05-02")
    # make_xlsx with closed_positions=None creates CLOSED sheet but writes no cells.
    make_xlsx(
        path,
        balance=Decimal("1"),
        equity=Decimal("2"),
        export_dt=datetime(2026, 5, 2, 10, 30),
        positions=[],
        closed_positions=None,
    )

    snapshot = _loader(input_dir).load_latest()
    assert snapshot.closed_positions == ()


def test_closed_sheet_with_partial_header_raises_malformed(tmp_path: Path) -> None:
    """CLOSED sheet has cells but is missing a required column → raises PortfolioMalformedError."""
    input_dir = tmp_path / "inputs"
    path = input_dir / _export_name("2026-05-02")
    make_xlsx(
        path,
        balance=Decimal("1"),
        equity=Decimal("2"),
        export_dt=datetime(2026, 5, 2, 10, 30),
        positions=[],
        closed_positions=[{"Position": 501}],
    )
    wb = load_workbook(path)
    ws = wb["CLOSED POSITION HISTORY"]
    gross_pl_col = _CLOSED_HEADER.index("Gross P/L") + 1
    ws.cell(row=_CLOSED_HEADER_ROW, column=gross_pl_col, value="REMOVED_COLUMN")
    wb.save(path)

    with pytest.raises(PortfolioMalformedError, match="gross_pl"):
        _loader(input_dir).load_latest()


def test_closed_sheet_with_renamed_column_raises_malformed(tmp_path: Path) -> None:
    """CLOSED sheet with all columns present but one renamed → raises PortfolioMalformedError."""
    input_dir = tmp_path / "inputs"
    path = input_dir / _export_name("2026-05-02")
    make_xlsx(
        path,
        balance=Decimal("1"),
        equity=Decimal("2"),
        export_dt=datetime(2026, 5, 2, 10, 30),
        positions=[],
        closed_positions=[{"Position": 501}],
    )
    wb = load_workbook(path)
    ws = wb["CLOSED POSITION HISTORY"]
    close_price_col = _CLOSED_HEADER.index("Close price") + 1
    ws.cell(row=_CLOSED_HEADER_ROW, column=close_price_col, value="Closing price")
    wb.save(path)

    with pytest.raises(PortfolioMalformedError, match="close_price"):
        _loader(input_dir).load_latest()


def test_closed_position_with_zero_open_price_raises_with_row_context(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    _write_export(input_dir, closed_positions=[{"Open price": Decimal("0")}])

    with pytest.raises(PortfolioMalformedError, match=r"closed-position row 12.*open_price"):
        _loader(input_dir).load_latest()


def test_closed_position_with_zero_close_price_raises_with_row_context(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    _write_export(input_dir, closed_positions=[{"Close price": Decimal("0")}])

    with pytest.raises(PortfolioMalformedError, match=r"closed-position row 12.*close_price"):
        _loader(input_dir).load_latest()


def test_position_with_zero_open_price_raises_with_row_context(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    _write_export(input_dir, positions=[{"Open price": Decimal("0")}])

    with pytest.raises(PortfolioMalformedError, match=r"open-position row 8.*open_price"):
        _loader(input_dir).load_latest()


def test_position_with_zero_market_price_raises_with_row_context(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs"
    _write_export(input_dir, positions=[{"Market price": Decimal("0")}])

    with pytest.raises(PortfolioMalformedError, match=r"open-position row 8.*market_price"):
        _loader(input_dir).load_latest()


def test_workbook_without_closed_sheet_yields_empty_closed_positions(tmp_path: Path) -> None:
    """Workbook with no CLOSED POSITION HISTORY sheet at all → closed_positions = ()."""
    input_dir = tmp_path / "inputs"
    path = input_dir / _export_name("2026-05-02")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default = wb.active
    assert default is not None
    wb.remove(default)
    ws = wb.create_sheet("OPEN POSITION 02052026")
    export_dt = datetime(2026, 5, 2, 10, 30)
    ws.cell(row=3, column=2, value=export_dt)
    ws.cell(row=4, column=4, value="Balance")
    ws.cell(row=4, column=7, value="Equity")
    ws.cell(row=5, column=4, value=1000.0)
    ws.cell(row=5, column=7, value=2000.0)
    for col, name in enumerate(_HEADER, start=1):
        ws.cell(row=7, column=col, value=name)
    ws.cell(row=8, column=1, value="Total")
    wb.save(path)

    snapshot = _loader(input_dir).load_latest()
    assert snapshot.closed_positions == ()
