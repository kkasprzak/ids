import logging
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ids.adapters.xtb_filename import parse_xtb_account_id, parse_xtb_filename
from ids.domain.enums import PositionType
from ids.domain.models import AccountSummary, PortfolioSnapshot, Position
from ids.domain.ports.portfolio import (
    NoPortfolioAvailableError,
    PortfolioLoader,
    PortfolioMalformedError,
)
from ids.domain.timezones import WARSAW

log = logging.getLogger(__name__)

_OPEN_SHEET_PREFIX = "OPEN POSITION "

# Logical field → accepted label aliases. Normalized before matching (strip + casefold).
# All XTB adapter complexity stays here; domain models never see these strings.
_POSITION_COLUMN_SCHEMA: dict[str, frozenset[str]] = {
    "position_id": frozenset({"Position"}),
    "symbol": frozenset({"Symbol"}),
    "type": frozenset({"Type"}),
    "volume": frozenset({"Volume"}),
    "open_time": frozenset({"Open time"}),
    "open_price": frozenset({"Open price"}),
    "market_price": frozenset({"Market price"}),
    "purchase_value": frozenset({"Purchase value"}),
    "sl": frozenset({"SL"}),
    "gross_pl": frozenset({"Gross P/L", "Gross P&L"}),
}
_ACCOUNT_COLUMN_SCHEMA: dict[str, frozenset[str]] = {
    "balance": frozenset({"Balance"}),
    "equity": frozenset({"Equity"}),
}

# Bounded scan caps: high enough to survive XTB layout changes, low enough to
# fail fast on corrupted or unrelated workbooks. Not domain rules.
_EXPORT_DATETIME_SCAN_ROWS = 15
_ACCOUNT_HEADER_SCAN_ROWS = 15
_POSITION_HEADER_SCAN_ROWS = 25


def _normalize_label(label: str) -> str:
    """Strip, collapse internal whitespace, and casefold for alias matching."""
    return " ".join(label.split()).casefold()


class XTBPortfolioLoader(PortfolioLoader):
    def __init__(self, input_dir: Path, ikze_account_id: str) -> None:
        self._input_dir = input_dir
        self._account_id = ikze_account_id

    def load_latest(self) -> PortfolioSnapshot:
        if not self._input_dir.is_dir():
            raise NoPortfolioAvailableError(
                f"Directory `{self._input_dir}/` not found.\n"
                f"  Create it and place an XTB XLSX export there:\n"
                f"    mkdir -p {self._input_dir}/\n"
                f"    cp ~/Downloads/account_ikze_{self._account_id}_*.xlsx {self._input_dir}/"
            )

        xlsx_files = sorted(self._input_dir.glob("*.xlsx"))
        candidates = [
            (parse_xtb_filename(path.name, expected_account_id=self._account_id), path)
            for path in xlsx_files
        ]
        matching = [(as_of, path) for as_of, path in candidates if as_of is not None]
        if matching:
            for as_of, path in candidates:
                if as_of is None:
                    log.debug("Skipped %s (does not match IKZE pattern)", path.name)
            matching.sort(key=lambda item: item[0])
            as_of_date, latest_file = matching[-1]
            return self._parse(latest_file, as_of_date)

        fallback_candidates: list[Path] = []
        excluded_other_accounts: list[str] = []
        for path in xlsx_files:
            declared_account_id = parse_xtb_account_id(path.name)
            if declared_account_id is not None and declared_account_id != self._account_id:
                excluded_other_accounts.append(path.name)
                continue
            fallback_candidates.append(path)

        if not fallback_candidates:
            found = [path.name for path in xlsx_files]
            excluded = excluded_other_accounts if excluded_other_accounts else "(none)"
            raise NoPortfolioAvailableError(
                f"No IKZE export matching pattern in `{self._input_dir}/` "
                f"and no eligible fallback XLSX files.\n"
                f"  Expected pattern:\n"
                f"    account_ikze_{self._account_id}_pl_xlsx_<YYYY-MM-DD>_<YYYY-MM-DD>.xlsx\n"
                f"  Found: {found if found else '(empty)'}\n"
                f"  Excluded as other account IDs: {excluded}"
            )

        latest_file = max(fallback_candidates, key=lambda path: path.stat().st_mtime)
        as_of_date = self._as_of_date_from_export_datetime(latest_file)
        log.info(
            "No strict IKZE filename matches; selected latest XLSX by mtime: %s "
            "(as_of_date from workbook export datetime: %s)",
            latest_file.name,
            as_of_date.isoformat(),
        )
        return self._parse(latest_file, as_of_date)

    def load_from_path(self, path: Path) -> PortfolioSnapshot:
        as_of = parse_xtb_filename(path.name, expected_account_id=self._account_id)
        if as_of is not None:
            return self._parse(path, as_of)

        declared_account_id = parse_xtb_account_id(path.name)
        if declared_account_id is not None and declared_account_id != self._account_id:
            raise PortfolioMalformedError(
                f"File `{path.name}` clearly belongs to IKZE account `{declared_account_id}`, "
                f"expected `{self._account_id}`."
            )
        as_of_from_workbook = self._as_of_date_from_export_datetime(path)
        log.info(
            "File %s does not match strict IKZE filename pattern; "
            "using workbook export datetime for as_of_date: %s",
            path.name,
            as_of_from_workbook.isoformat(),
        )
        return self._parse(path, as_of_from_workbook)

    def _parse(self, path: Path, as_of_date: date) -> PortfolioSnapshot:
        try:
            workbook = load_workbook(path, data_only=True, read_only=True)
            sheet = self._find_open_position_sheet(workbook)
            account = self._parse_account_summary(sheet)
            positions = self._parse_positions(sheet)
        except PortfolioMalformedError:
            raise
        except Exception as exc:
            raise PortfolioMalformedError(f"Failed to parse {path.name}: {exc}") from exc

        return PortfolioSnapshot(
            as_of_date=as_of_date,
            source_id=f"xtb:{path.name}",
            account=account,
            positions=tuple(positions),
        )

    def _find_open_position_sheet(self, workbook: Workbook) -> Worksheet:
        prefix_matches = [n for n in workbook.sheetnames if n.startswith(_OPEN_SHEET_PREFIX)]
        if prefix_matches:
            return workbook[prefix_matches[0]]

        semantic_matches = [
            n for n in workbook.sheetnames if _sheet_has_position_table(workbook[n])
        ]
        if len(semantic_matches) == 1:
            log.info(
                "No sheet matching prefix %r; using semantic fallback: %r",
                _OPEN_SHEET_PREFIX,
                semantic_matches[0],
            )
            return workbook[semantic_matches[0]]
        if len(semantic_matches) > 1:
            raise PortfolioMalformedError(
                f"Multiple sheets contain the open-position table columns; cannot choose "
                f"automatically. Ambiguous sheets: {semantic_matches}. "
                f"All sheets: {workbook.sheetnames}"
            )
        required_fields = sorted(_POSITION_COLUMN_SCHEMA.keys())
        raise PortfolioMalformedError(
            f"No sheet starting with `{_OPEN_SHEET_PREFIX}` found and no sheet contains "
            f"the required open-position fields {required_fields}. "
            f"Sheets: {workbook.sheetnames}"
        )

    def _parse_account_summary(self, sheet: Worksheet) -> AccountSummary:
        label_row_idx, label_columns = _find_labelled_row(
            sheet, schema=_ACCOUNT_COLUMN_SCHEMA, max_scan_rows=_ACCOUNT_HEADER_SCAN_ROWS
        )
        value_row = next(
            sheet.iter_rows(
                min_row=label_row_idx + 1,
                max_row=label_row_idx + 1,
                values_only=True,
            )
        )
        balance = _cell_to_decimal(value_row[label_columns["balance"]])
        equity = _cell_to_decimal(value_row[label_columns["equity"]])
        export_dt = _find_export_datetime(sheet)
        return AccountSummary(balance_pln=balance, equity_pln=equity, export_datetime=export_dt)

    def _parse_positions(self, sheet: Worksheet) -> list[Position]:
        header_row_idx, columns = _find_labelled_row(
            sheet, schema=_POSITION_COLUMN_SCHEMA, max_scan_rows=_POSITION_HEADER_SCAN_ROWS
        )

        # Footer detection: any non-numeric value in the position_id cell ends the table.
        # This covers "Total", "TOTAL", localized variants, and future XTB footer changes
        # without requiring a hardcoded alias list. Empty cells (None) are spacer rows.
        # Numeric rows that have bad values in other fields still raise PortfolioMalformedError.
        positions: list[Position] = []
        for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
            first = row[columns["position_id"]] if columns["position_id"] < len(row) else None
            if first is None:
                continue
            if not isinstance(first, int | float):
                break
            positions.append(self._row_to_position(row, columns))
        return positions

    def _row_to_position(self, row: tuple[Any, ...], columns: dict[str, int]) -> Position:
        sl_raw = row[columns["sl"]]
        sl = None if sl_raw in (None, 0, 0.0) else _cell_to_decimal(sl_raw)
        return Position(
            id=int(row[columns["position_id"]]),
            symbol=str(row[columns["symbol"]]),
            type=PositionType(row[columns["type"]]),
            volume=_cell_to_decimal(row[columns["volume"]]),
            open_time=_naive_dt_to_warsaw(row[columns["open_time"]]),
            open_price=_cell_to_decimal(row[columns["open_price"]]),
            market_price=_cell_to_decimal(row[columns["market_price"]]),
            purchase_value_pln=_cell_to_decimal(row[columns["purchase_value"]]),
            gross_pl_pln=_cell_to_decimal(row[columns["gross_pl"]]),
            sl=sl,
        )

    def _as_of_date_from_export_datetime(self, path: Path) -> date:
        message_prefix = (
            f"Could not derive as_of_date for `{path.name}` from workbook export datetime"
        )
        try:
            workbook = load_workbook(path, data_only=True, read_only=True)
            sheet = self._find_open_position_sheet(workbook)
            return _find_export_datetime(sheet).date()
        except PortfolioMalformedError as exc:
            raise PortfolioMalformedError(f"{message_prefix}: {exc}") from exc
        except Exception as exc:
            raise PortfolioMalformedError(f"{message_prefix}: {exc}") from exc


def _sheet_has_position_table(
    sheet: Worksheet, max_scan_rows: int = _POSITION_HEADER_SCAN_ROWS
) -> bool:
    required_count = len(_POSITION_COLUMN_SCHEMA)
    for row in sheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True):
        present = {_normalize_label(str(cell)) for cell in row if isinstance(cell, str)}
        # Count distinct logical fields covered (each field matches if any alias is present)
        covered = sum(
            1
            for aliases in _POSITION_COLUMN_SCHEMA.values()
            if any(_normalize_label(a) in present for a in aliases)
        )
        if covered == required_count:
            return True
    return False


def _find_labelled_row(
    sheet: Worksheet,
    *,
    schema: dict[str, frozenset[str]],
    max_scan_rows: int,
) -> tuple[int, dict[str, int]]:
    normalized_schema = {
        logical: {_normalize_label(a) for a in aliases} for logical, aliases in schema.items()
    }
    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1
    ):
        norm_row = {
            _normalize_label(str(cell)): idx
            for idx, cell in enumerate(row)
            if isinstance(cell, str)
        }
        columns: dict[str, int] = {}
        for logical, norm_aliases in normalized_schema.items():
            for alias in norm_aliases:
                if alias in norm_row:
                    columns[logical] = norm_row[alias]
                    break
        if len(columns) == len(schema):
            return row_idx, columns
    required_desc = [
        f"{logical} (accepts: {', '.join(sorted(aliases))})" for logical, aliases in schema.items()
    ]
    raise PortfolioMalformedError(
        f"Could not find row with all required columns in first {max_scan_rows} rows. "
        f"Required: {'; '.join(required_desc)}"
    )


def _cell_to_decimal(value: Any) -> Decimal:
    if value is None:
        raise PortfolioMalformedError("Expected number, got None")
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _naive_dt_to_warsaw(value: Any) -> datetime:
    if not isinstance(value, datetime):
        raise PortfolioMalformedError(f"Expected datetime, got {type(value).__name__}")
    return value.replace(tzinfo=WARSAW) if value.tzinfo is None else value.astimezone(WARSAW)


def _find_export_datetime(sheet: Worksheet) -> datetime:
    for row in sheet.iter_rows(min_row=1, max_row=_EXPORT_DATETIME_SCAN_ROWS, values_only=True):
        for cell in row:
            if isinstance(cell, datetime):
                return (
                    cell.replace(tzinfo=WARSAW) if cell.tzinfo is None else cell.astimezone(WARSAW)
                )
    raise PortfolioMalformedError(
        f"Could not find export datetime in first {_EXPORT_DATETIME_SCAN_ROWS} rows."
    )
