import logging
from dataclasses import dataclass, replace
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import NoReturn

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ids.application.ports.portfolio import (
    NoPortfolioAvailableError,
    PortfolioLoader,
    PortfolioMalformedError,
)
from ids.domain.enums import PositionType
from ids.domain.models import AccountSummary, ClosedPosition, PortfolioSnapshot, Position
from ids.domain.timezones import WARSAW
from ids.domain.value_objects import Price, Symbol
from ids.infrastructure.adapters.xtb_filename import parse_xtb_account_id, parse_xtb_filename

log = logging.getLogger(__name__)

_OPEN_SHEET_PREFIX = "OPEN POSITION "
_CLOSED_SHEET_NAME = "CLOSED POSITION HISTORY"

type RawCellValue = object
type RawDataRow = tuple[RawCellValue, ...]
type ColumnIndexes = dict[str, int]


@dataclass(frozen=True)
class _XTBLabel:
    """Canonical XTB workbook label for case-insensitive, whitespace-tolerant matching."""

    value: str

    def __post_init__(self) -> None:
        object.__setattr__(self, "value", " ".join(self.value.split()).lower())

    def __str__(self) -> str:
        return self.value


type ColumnSchema = dict[str, frozenset[_XTBLabel]]


def _labels(*values: str) -> frozenset[_XTBLabel]:
    return frozenset(_XTBLabel(value) for value in values)


# Logical field → accepted label aliases.
# All XTB adapter complexity stays here; domain models never see these strings.
_POSITION_COLUMN_SCHEMA: ColumnSchema = {
    "position_id": _labels("Position"),
    "symbol": _labels("Symbol"),
    "type": _labels("Type"),
    "volume": _labels("Volume"),
    "open_time": _labels("Open time"),
    "open_price": _labels("Open price"),
    "market_price": _labels("Market price"),
    "purchase_value": _labels("Purchase value"),
    "sl": _labels("SL"),
    "gross_pl": _labels("Gross P/L", "Gross P&L"),
}
_CLOSED_POSITION_COLUMN_SCHEMA: ColumnSchema = {
    "position_id": _labels("Position"),
    "symbol": _labels("Symbol"),
    "type": _labels("Type"),
    "volume": _labels("Volume"),
    "open_time": _labels("Open time"),
    "open_price": _labels("Open price"),
    "close_time": _labels("Close time"),
    "close_price": _labels("Close price"),
    "purchase_value": _labels("Purchase value"),
    "gross_pl": _labels("Gross P/L", "Gross P&L"),
}
_ACCOUNT_COLUMN_SCHEMA: ColumnSchema = {
    "balance": _labels("Balance"),
    "equity": _labels("Equity"),
}

# Bounded scan caps: high enough to survive XTB layout changes, low enough to
# fail fast on corrupted or unrelated workbooks. Not domain rules.
_EXPORT_DATETIME_SCAN_ROWS = 15
_ACCOUNT_HEADER_SCAN_ROWS = 15
_POSITION_HEADER_SCAN_ROWS = 25
# Real fixture has the closed-position header at row 11; 15 gives headroom.
_CLOSED_POSITION_HEADER_SCAN_ROWS = 15


class XTBPortfolioLoader(PortfolioLoader):
    def __init__(self, input_dir: Path, ikze_account_id: str) -> None:
        self._input_dir = input_dir
        self._account_id = ikze_account_id

    def load_latest(self) -> PortfolioSnapshot:
        latest_file, as_of_date = self._select_latest_export()
        return self._parse(latest_file, as_of_date)

    def load_from_path(self, path: Path) -> PortfolioSnapshot:
        as_of_date = self._as_of_date_for_path(path)
        return self._parse(path, as_of_date)

    def _select_latest_export(self) -> tuple[Path, date]:
        xlsx_files = self._list_xlsx_files()

        strict_match = self._latest_strict_filename_match(xlsx_files)
        if strict_match is not None:
            return strict_match

        fallback_candidates, excluded_other_accounts = self._eligible_fallback_candidates(
            xlsx_files
        )
        if not fallback_candidates:
            self._raise_no_available_export(xlsx_files, excluded_other_accounts)

        latest_file = max(fallback_candidates, key=self._mtime_for_path)
        as_of_date = self._as_of_date_from_export_datetime(latest_file)
        log.info(
            "No strict IKZE filename matches; selected latest XLSX by mtime: %s "
            "(as_of_date from workbook export datetime: %s)",
            latest_file.name,
            as_of_date.isoformat(),
        )
        return latest_file, as_of_date

    def _list_xlsx_files(self) -> list[Path]:
        try:
            if not self._input_dir.is_dir():
                raise NoPortfolioAvailableError(
                    f"Directory `{self._input_dir}/` not found.\n"
                    f"  Create it and place an XTB XLSX export there:\n"
                    f"    mkdir -p {self._input_dir}/\n"
                    f"    cp ~/Downloads/account_ikze_{self._account_id}_*.xlsx {self._input_dir}/"
                )
            return sorted(self._input_dir.glob("*.xlsx"))
        except NoPortfolioAvailableError:
            raise
        except OSError as exc:
            raise NoPortfolioAvailableError(
                f"Could not access directory `{self._input_dir}/`: {exc}"
            ) from exc

    def _latest_strict_filename_match(self, xlsx_files: list[Path]) -> tuple[Path, date] | None:
        candidates = [
            (parse_xtb_filename(path.name, expected_account_id=self._account_id), path)
            for path in xlsx_files
        ]
        matching = [(as_of, path) for as_of, path in candidates if as_of is not None]
        if not matching:
            return None

        for as_of, path in candidates:
            if as_of is None:
                log.debug("Skipped %s (does not match IKZE pattern)", path.name)
        as_of_date, latest_file = max(matching, key=lambda item: item[0])
        return latest_file, as_of_date

    def _eligible_fallback_candidates(self, xlsx_files: list[Path]) -> tuple[list[Path], list[str]]:
        fallback_candidates: list[Path] = []
        excluded_other_accounts: list[str] = []
        for path in xlsx_files:
            declared_account_id = parse_xtb_account_id(path.name)
            if declared_account_id is not None and declared_account_id != self._account_id:
                excluded_other_accounts.append(path.name)
                continue
            fallback_candidates.append(path)
        return fallback_candidates, excluded_other_accounts

    def _raise_no_available_export(
        self, xlsx_files: list[Path], excluded_other_accounts: list[str]
    ) -> NoReturn:
        found = [path.name for path in xlsx_files]
        excluded = excluded_other_accounts if excluded_other_accounts else "(none)"
        raise NoPortfolioAvailableError(
            f"No IKZE export matching pattern in `{self._input_dir}/` "
            f"and no eligible fallback XLSX files.\n"
            f"  Expected pattern:\n"
            f"    account_ikze_{self._account_id}_pl_xlsx_<YYYY-MM-DD>_<YYYY-MM-DD>.xlsx\n"
            f"  Found: {found if found else '(empty)'}\n"
            f"  Excluded as other account IDs: {excluded}"
        )

    def _mtime_for_path(self, path: Path) -> float:
        try:
            return path.stat().st_mtime
        except OSError as exc:
            raise PortfolioMalformedError(
                f"Failed to inspect XLSX file `{path.name}`: {exc}"
            ) from exc

    def _as_of_date_for_path(self, path: Path) -> date:
        as_of = parse_xtb_filename(path.name, expected_account_id=self._account_id)
        if as_of is not None:
            return as_of

        declared_account_id = parse_xtb_account_id(path.name)
        if declared_account_id is not None and declared_account_id != self._account_id:
            raise PortfolioMalformedError(
                f"File `{path.name}` clearly belongs to IKZE account `{declared_account_id}`, "
                f"expected `{self._account_id}`."
            )
        as_of_from_workbook = self._as_of_date_from_export_datetime(path)
        log.info(
            "File %s does not match strict IKZE filename pattern; "
            "using workbook export datetime for as_of_date: %s",
            path.name,
            as_of_from_workbook.isoformat(),
        )
        return as_of_from_workbook

    def _parse(self, path: Path, as_of_date: date) -> PortfolioSnapshot:
        try:
            workbook = load_workbook(path, data_only=True, read_only=True)
            open_sheet = self._find_open_position_sheet(workbook)
            account = self._parse_account_summary(open_sheet)
            positions = self._parse_positions(open_sheet)
            closed_positions = self._parse_closed_positions(workbook)
        except PortfolioMalformedError:
            raise
        except Exception as exc:
            raise PortfolioMalformedError(f"Failed to parse {path.name}: {exc}") from exc

        return PortfolioSnapshot(
            as_of_date=as_of_date,
            source_id=f"xtb:{path.name}",
            account=account,
            positions=tuple(positions),
            closed_positions=tuple(closed_positions),
        )

    def _load_open_position_sheet(self, path: Path) -> Worksheet:
        workbook = load_workbook(path, data_only=True, read_only=True)
        return self._find_open_position_sheet(workbook)

    def _find_open_position_sheet(self, workbook: Workbook) -> Worksheet:
        prefix_matches = [n for n in workbook.sheetnames if n.startswith(_OPEN_SHEET_PREFIX)]
        if prefix_matches:
            return workbook[prefix_matches[0]]

        semantic_matches = [
            n for n in workbook.sheetnames if _sheet_has_position_table(workbook[n])
        ]
        if len(semantic_matches) == 1:
            log.info(
                "No sheet matching prefix %r; using semantic fallback: %r",
                _OPEN_SHEET_PREFIX,
                semantic_matches[0],
            )
            return workbook[semantic_matches[0]]
        if len(semantic_matches) > 1:
            raise PortfolioMalformedError(
                f"Multiple sheets contain the open-position table columns; cannot choose "
                f"automatically. Ambiguous sheets: {semantic_matches}. "
                f"All sheets: {workbook.sheetnames}"
            )
        required_fields = sorted(_POSITION_COLUMN_SCHEMA.keys())
        raise PortfolioMalformedError(
            f"No sheet starting with `{_OPEN_SHEET_PREFIX}` found and no sheet contains "
            f"the required open-position fields {required_fields}. "
            f"Sheets: {workbook.sheetnames}"
        )

    def _parse_account_summary(self, sheet: Worksheet) -> AccountSummary:
        label_row_idx, label_columns = _find_labelled_row(
            sheet, schema=_ACCOUNT_COLUMN_SCHEMA, max_scan_rows=_ACCOUNT_HEADER_SCAN_ROWS
        )
        value_row = next(
            sheet.iter_rows(
                min_row=label_row_idx + 1,
                max_row=label_row_idx + 1,
                values_only=True,
            )
        )
        balance = _cell_to_decimal(value_row[label_columns["balance"]])
        equity = _cell_to_decimal(value_row[label_columns["equity"]])
        export_dt = _find_export_datetime(sheet)
        return AccountSummary(balance_pln=balance, equity_pln=equity, export_datetime=export_dt)

    def _parse_positions(self, sheet: Worksheet) -> list[Position]:
        header_row_idx, columns = _find_labelled_row(
            sheet, schema=_POSITION_COLUMN_SCHEMA, max_scan_rows=_POSITION_HEADER_SCAN_ROWS
        )

        # Footer detection: any non-numeric value in the position_id cell ends the table.
        # This covers "Total", "TOTAL", localized variants, and future XTB footer changes
        # without requiring a hardcoded alias list. Empty cells (None) are spacer rows.
        # Numeric rows that have bad values in other fields still raise PortfolioMalformedError.
        positions: list[Position] = []
        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=header_row_idx + 1, values_only=True),
            start=header_row_idx + 1,
        ):
            first = row[columns["position_id"]] if columns["position_id"] < len(row) else None
            if first is None:
                continue
            if not isinstance(first, int | float):
                break
            positions.append(self._row_to_position(row, columns, row_idx))
        return positions

    def _row_to_position(self, row: RawDataRow, columns: ColumnIndexes, row_idx: int) -> Position:
        return Position(
            id=_row_id(row_idx, self._row_cell(row, columns, row_idx, "position_id")),
            symbol=_row_symbol(row_idx, self._row_cell(row, columns, row_idx, "symbol")),
            type=_row_type(row_idx, self._row_cell(row, columns, row_idx, "type")),
            volume=_row_decimal(row_idx, "volume", self._row_cell(row, columns, row_idx, "volume")),
            open_time=_row_datetime(
                row_idx, "open_time", self._row_cell(row, columns, row_idx, "open_time")
            ),
            open_price=_row_price(
                row_idx, "open_price", self._row_cell(row, columns, row_idx, "open_price")
            ),
            market_price=_row_price(
                row_idx, "market_price", self._row_cell(row, columns, row_idx, "market_price")
            ),
            purchase_value_pln=_row_decimal(
                row_idx,
                "purchase_value",
                self._row_cell(row, columns, row_idx, "purchase_value"),
            ),
            gross_pl_pln=_row_decimal(
                row_idx, "gross_pl", self._row_cell(row, columns, row_idx, "gross_pl")
            ),
            sl=_row_stop_loss(row_idx, self._row_cell(row, columns, row_idx, "sl")),
        )

    def _row_cell(
        self, row: RawDataRow, columns: ColumnIndexes, row_idx: int, logical_field: str
    ) -> RawCellValue:
        col_idx = columns[logical_field]
        if col_idx >= len(row):
            raise PortfolioMalformedError(
                f"Malformed open-position row {row_idx}: missing field `{logical_field}` "
                f"at expected column index {col_idx}."
            )
        return row[col_idx]

    def _find_closed_position_sheet(self, workbook: Workbook) -> Worksheet | None:
        # 1. Exact match.
        if _CLOSED_SHEET_NAME in workbook.sheetnames:
            return workbook[_CLOSED_SHEET_NAME]
        # 2. Normalized match (case-insensitive, whitespace-tolerant).
        target = _XTBLabel(_CLOSED_SHEET_NAME)
        for name in workbook.sheetnames:
            if _XTBLabel(name) == target:
                log.info("Closed-position sheet found via normalized match: %r", name)
                return workbook[name]
        # 3. Semantic fallback by column structure.
        semantic = [n for n in workbook.sheetnames if _sheet_has_closed_position_table(workbook[n])]
        if len(semantic) == 1:
            log.info("No closed-sheet name match; using semantic fallback: %r", semantic[0])
            return workbook[semantic[0]]
        if len(semantic) > 1:
            raise PortfolioMalformedError(
                f"Multiple sheets contain the closed-position table columns; cannot choose "
                f"automatically. Ambiguous sheets: {semantic}. "
                f"All sheets: {workbook.sheetnames}"
            )
        return None

    def _parse_closed_positions(self, workbook: Workbook) -> list[ClosedPosition]:
        sheet = self._find_closed_position_sheet(workbook)
        if sheet is None:
            return []
        has_cells = any(
            cell is not None
            for row in sheet.iter_rows(
                min_row=1, max_row=_CLOSED_POSITION_HEADER_SCAN_ROWS, values_only=True
            )
            for cell in row
        )
        if not has_cells:
            return []
        header_row_idx, columns = _find_labelled_row(
            sheet,
            schema=_CLOSED_POSITION_COLUMN_SCHEMA,
            max_scan_rows=_CLOSED_POSITION_HEADER_SCAN_ROWS,
        )

        closed_positions: list[ClosedPosition] = []
        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=header_row_idx + 1, values_only=True),
            start=header_row_idx + 1,
        ):
            first = row[columns["position_id"]] if columns["position_id"] < len(row) else None
            if first is None:
                continue
            if not isinstance(first, int | float):
                break
            closed_positions.append(self._row_to_closed_position(row, columns, row_idx))
        return _aggregate_closed_by_id(closed_positions)

    def _row_to_closed_position(
        self, row: RawDataRow, columns: ColumnIndexes, row_idx: int
    ) -> ClosedPosition:
        return ClosedPosition(
            id=_row_id(row_idx, self._closed_row_cell(row, columns, row_idx, "position_id")),
            symbol=_row_symbol(row_idx, self._closed_row_cell(row, columns, row_idx, "symbol")),
            type=_row_type(row_idx, self._closed_row_cell(row, columns, row_idx, "type")),
            volume=_row_decimal(
                row_idx, "volume", self._closed_row_cell(row, columns, row_idx, "volume")
            ),
            open_time=_row_datetime(
                row_idx, "open_time", self._closed_row_cell(row, columns, row_idx, "open_time")
            ),
            close_time=_row_datetime(
                row_idx, "close_time", self._closed_row_cell(row, columns, row_idx, "close_time")
            ),
            open_price=_row_price(
                row_idx,
                "open_price",
                self._closed_row_cell(row, columns, row_idx, "open_price"),
            ),
            close_price=_row_price(
                row_idx,
                "close_price",
                self._closed_row_cell(row, columns, row_idx, "close_price"),
            ),
            purchase_value_pln=_row_decimal(
                row_idx,
                "purchase_value",
                self._closed_row_cell(row, columns, row_idx, "purchase_value"),
            ),
            gross_pl_pln=_row_decimal(
                row_idx, "gross_pl", self._closed_row_cell(row, columns, row_idx, "gross_pl")
            ),
        )

    def _closed_row_cell(
        self, row: RawDataRow, columns: ColumnIndexes, row_idx: int, logical_field: str
    ) -> RawCellValue:
        col_idx = columns[logical_field]
        if col_idx >= len(row):
            raise PortfolioMalformedError(
                f"Malformed closed-position row {row_idx}: missing field `{logical_field}` "
                f"at expected column index {col_idx}."
            )
        return row[col_idx]

    def _as_of_date_from_export_datetime(self, path: Path) -> date:
        message_prefix = (
            f"Could not derive as_of_date for `{path.name}` from workbook export datetime"
        )
        try:
            sheet = self._load_open_position_sheet(path)
            return _find_export_datetime(sheet).date()
        except PortfolioMalformedError as exc:
            raise PortfolioMalformedError(f"{message_prefix}: {exc}") from exc
        except Exception as exc:
            raise PortfolioMalformedError(f"{message_prefix}: {exc}") from exc


def _sheet_has_position_table(
    sheet: Worksheet, max_scan_rows: int = _POSITION_HEADER_SCAN_ROWS
) -> bool:
    required_count = len(_POSITION_COLUMN_SCHEMA)
    for row in sheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True):
        present = {_XTBLabel(str(cell)) for cell in row if isinstance(cell, str)}
        # Count distinct logical fields covered (each field matches if any alias is present)
        covered = sum(
            1
            for aliases in _POSITION_COLUMN_SCHEMA.values()
            if any(alias in present for alias in aliases)
        )
        if covered == required_count:
            return True
    return False


def _sheet_has_closed_position_table(
    sheet: Worksheet, max_scan_rows: int = _CLOSED_POSITION_HEADER_SCAN_ROWS
) -> bool:
    required_count = len(_CLOSED_POSITION_COLUMN_SCHEMA)
    for row in sheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True):
        present = {_XTBLabel(str(cell)) for cell in row if isinstance(cell, str)}
        covered = sum(
            1
            for aliases in _CLOSED_POSITION_COLUMN_SCHEMA.values()
            if any(alias in present for alias in aliases)
        )
        if covered == required_count:
            return True
    return False


def _aggregate_closed_by_id(positions: list[ClosedPosition]) -> list[ClosedPosition]:
    """Fold partial-fill rows sharing one XTB position id into one closed position.

    XTB reports a position closed in tranches as one row per fill — same id, open
    time and prices, differing only in volume, gross P/L and close time. Summing
    the fills reconstructs the single logical position; without this each fill
    would overwrite the last in the position log, losing realized P/L.
    """
    aggregated: dict[int, ClosedPosition] = {}
    for position in positions:
        existing = aggregated.get(position.id)
        if existing is None:
            aggregated[position.id] = position
            continue
        aggregated[position.id] = replace(
            existing,
            volume=existing.volume + position.volume,
            gross_pl_pln=existing.gross_pl_pln + position.gross_pl_pln,
            purchase_value_pln=existing.purchase_value_pln + position.purchase_value_pln,
            close_time=max(existing.close_time, position.close_time),
        )
    return list(aggregated.values())


def _find_labelled_row(
    sheet: Worksheet,
    *,
    schema: ColumnSchema,
    max_scan_rows: int,
) -> tuple[int, ColumnIndexes]:
    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1
    ):
        norm_row = {
            _XTBLabel(str(cell)): idx for idx, cell in enumerate(row) if isinstance(cell, str)
        }
        columns: ColumnIndexes = {}
        for logical, aliases in schema.items():
            for alias in aliases:
                if alias in norm_row:
                    columns[logical] = norm_row[alias]
                    break
        if len(columns) == len(schema):
            return row_idx, columns
    required_desc = [
        f"{logical} (accepts: {', '.join(sorted(str(alias) for alias in aliases))})"
        for logical, aliases in schema.items()
    ]
    raise PortfolioMalformedError(
        f"Could not find row with all required columns in first {max_scan_rows} rows. "
        f"Required: {'; '.join(required_desc)}"
    )


def _cell_to_decimal(value: RawCellValue) -> Decimal:
    if value is None:
        raise PortfolioMalformedError("Expected number, got None")
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _cell_to_symbol(value: RawCellValue) -> Symbol:
    if value is None:
        raise PortfolioMalformedError("Expected symbol, got None")
    return Symbol(str(value))


def _row_symbol(row_idx: int, value: RawCellValue) -> Symbol:
    try:
        return _cell_to_symbol(value)
    except Exception as exc:
        raise _row_error(row_idx, "symbol", value, exc) from exc


def _row_type(row_idx: int, value: RawCellValue) -> PositionType:
    try:
        return PositionType(value)
    except Exception as exc:
        raise _row_error(row_idx, "type", value, exc) from exc


def _row_decimal(row_idx: int, field_name: str, value: RawCellValue) -> Decimal:
    try:
        return _cell_to_decimal(value)
    except Exception as exc:
        raise _row_error(row_idx, field_name, value, exc) from exc


def _row_price(row_idx: int, field_name: str, value: RawCellValue) -> Price:
    try:
        return Price(_cell_to_decimal(value))
    except Exception as exc:
        raise _row_error(row_idx, field_name, value, exc) from exc


def _row_datetime(row_idx: int, field_name: str, value: RawCellValue) -> datetime:
    try:
        return _naive_dt_to_warsaw(value)
    except Exception as exc:
        raise _row_error(row_idx, field_name, value, exc) from exc


def _row_id(row_idx: int, value: RawCellValue) -> int:
    try:
        return _cell_to_int(value)
    except Exception as exc:
        raise _row_error(row_idx, "position_id", value, exc) from exc


def _row_stop_loss(row_idx: int, value: RawCellValue) -> Decimal | None:
    try:
        return _normalize_stop_loss(value)
    except Exception as exc:
        raise _row_error(row_idx, "sl", value, exc) from exc


def _cell_to_int(value: RawCellValue) -> int:
    if isinstance(value, str | int | float | Decimal):
        return int(value)
    raise PortfolioMalformedError(f"Expected integer-compatible value, got {type(value).__name__}")


def _naive_dt_to_warsaw(value: RawCellValue) -> datetime:
    if not isinstance(value, datetime):
        raise PortfolioMalformedError(f"Expected datetime, got {type(value).__name__}")
    return value.replace(tzinfo=WARSAW) if value.tzinfo is None else value.astimezone(WARSAW)


def _find_export_datetime(sheet: Worksheet) -> datetime:
    for row in sheet.iter_rows(min_row=1, max_row=_EXPORT_DATETIME_SCAN_ROWS, values_only=True):
        for cell in row:
            if isinstance(cell, datetime):
                return (
                    cell.replace(tzinfo=WARSAW) if cell.tzinfo is None else cell.astimezone(WARSAW)
                )
    raise PortfolioMalformedError(
        f"Could not find export datetime in first {_EXPORT_DATETIME_SCAN_ROWS} rows."
    )


def _row_error(
    row_idx: int, field_name: str, value: RawCellValue, exc: Exception
) -> PortfolioMalformedError:
    return PortfolioMalformedError(
        f"Malformed row {row_idx}, field `{field_name}` "
        f"(value={value!r}, type={type(value).__name__}): {exc}"
    )


def _normalize_stop_loss(value: RawCellValue) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == "":
            return None
        value = stripped
    decimal_value = _cell_to_decimal(value)
    return None if decimal_value == 0 else decimal_value
